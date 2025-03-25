import openpyxl
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def crear_excel_costos_kombucha():
    # Crear libro de trabajo
    wb = openpyxl.Workbook()
    
    # Eliminar hoja por defecto si existe
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 1. Hoja de Resumen (Principal)
    resumen = wb.create_sheet("Resumen", 0)
    
    # 2. Hojas de detalle
    produccion = wb.create_sheet("Costos Producción")
    logistica = wb.create_sheet("Costos Logística")
    operacion = wb.create_sheet("Costos Operación")
    ventas = wb.create_sheet("Costos Ventas")
    
    # Datos comunes
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    años = [2023, 2024]  # Datos para 2 años
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_aligned = Alignment(horizontal="center")
    
    # --- Hoja de Costos de Producción ---
    produccion.append(["Mes"] + meses)
    produccion.append(["Materia Prima"] + [4.2, 4.5, 4.8, 5.1, 5.3, 5.6, 5.9, 6.2, 6.0, 5.8, 5.5, 5.2])
    produccion.append(["Mano de Obra"] + [3.8, 3.9, 4.0, 4.1, 4.2, 4.3, 4.4, 4.5, 4.3, 4.2, 4.1, 4.0])
    produccion.append(["Envases"] + [1.5, 1.6, 1.7, 1.8, 1.9, 2.0, 2.1, 2.2, 2.0, 1.9, 1.8, 1.7])
    produccion.append(["Total Producción"] + [9.5, 10.0, 10.5, 11.0, 11.4, 11.9, 12.4, 12.9, 12.3, 11.9, 11.4, 10.9])
    
    # Aplicar formato
    for row in produccion.iter_rows(max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_aligned
    
    # --- Hoja de Costos Logística ---
    logistica.append(["Mes"] + meses)
    logistica.append(["Transporte"] + [1.2, 1.3, 1.4, 1.5, 1.6, 1.7, 1.8, 1.9, 1.7, 1.6, 1.5, 1.4])
    logistica.append(["Almacenamiento"] + [0.8, 0.8, 0.9, 0.9, 1.0, 1.0, 1.1, 1.1, 1.0, 1.0, 0.9, 0.9])
    logistica.append(["Total Logística"] + [2.0, 2.1, 2.3, 2.4, 2.6, 2.7, 2.9, 3.0, 2.7, 2.6, 2.4, 2.3])
    
    for row in logistica.iter_rows(max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_aligned
    
    # --- Hoja de Costos Operación ---
    operacion.append(["Mes"] + meses)
    operacion.append(["Energía"] + [0.9, 0.9, 0.8, 0.8, 0.9, 1.0, 1.1, 1.1, 1.0, 0.9, 0.9, 0.9])
    operacion.append(["Agua"] + [0.3, 0.3, 0.3, 0.3, 0.4, 0.4, 0.4, 0.4, 0.4, 0.3, 0.3, 0.3])
    operacion.append(["Mantenimiento"] + [0.5, 0.5, 0.6, 0.6, 0.6, 0.7, 0.7, 0.7, 0.6, 0.6, 0.5, 0.5])
    operacion.append(["Total Operación"] + [1.7, 1.7, 1.7, 1.7, 1.9, 2.1, 2.2, 2.2, 2.0, 1.8, 1.7, 1.7])
    
    for row in operacion.iter_rows(max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_aligned
    
    # --- Hoja de Costos Ventas ---
    ventas.append(["Mes"] + meses)
    ventas.append(["Marketing"] + [1.5, 1.6, 1.8, 2.0, 2.2, 2.4, 2.6, 2.8, 2.5, 2.3, 2.0, 1.8])
    ventas.append(["Comisiones"] + [0.8, 0.9, 1.0, 1.1, 1.2, 1.3, 1.4, 1.5, 1.3, 1.2, 1.1, 1.0])
    ventas.append(["Total Ventas"] + [2.3, 2.5, 2.8, 3.1, 3.4, 3.7, 4.0, 4.3, 3.8, 3.5, 3.1, 2.8])
    
    for row in ventas.iter_rows(max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_aligned
    
    # --- Hoja Resumen ---
    # Encabezados
    resumen.append(["Resumen de Costos Anuales - Fábrica de Kombucha"])
    resumen.append(["Mes", "Producción", "Logística", "Operación", "Ventas", "Total"])
    
    # Obtener totales de cada hoja
    for i, mes in enumerate(meses, start=2):  # Empezar en 2 por el encabezado
        prod = produccion.cell(row=5, column=i).value  # Total Producción
        log = logistica.cell(row=4, column=i).value   # Total Logística
        op = operacion.cell(row=5, column=i).value    # Total Operación
        vta = ventas.cell(row=4, column=i).value      # Total Ventas
        total = prod + log + op + vta
        
        resumen.append([mes, prod, log, op, vta, total])
    
    # Formato hoja resumen
    for row in resumen.iter_rows(max_row=2):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_aligned
    
    # Ajustar anchos de columna
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
    
    # --- Crear gráfico en hoja Resumen ---
    chart = BarChart()
    chart.title = "Costos por Área - Fábrica de Kombucha"
    chart.y_axis.title = "Costos (miles USD)"
    chart.x_axis.title = "Meses"
    chart.style = 10
    
    # Referencias de datos
    data = Reference(resumen, min_col=2, max_col=5, min_row=2, max_row=13)
    cats = Reference(resumen, min_col=1, min_row=3, max_row=13)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # Gráfico de línea para el total
    line_chart = LineChart()
    line_chart.title = "Costo Total Mensual"
    line_data = Reference(resumen, min_col=6, min_row=2, max_row=13)
    line_chart.add_data(line_data, titles_from_data=True)
    
    # Posicionar gráficos
    resumen.add_chart(chart, "H2")
    resumen.add_chart(line_chart, "H20")
    
    # Guardar archivo
    nombre_archivo = "Costos_Fabrica_Kombucha.xlsx"
    wb.save(nombre_archivo)
    print(f"Archivo '{nombre_archivo}' creado exitosamente!")

# Ejecutar la función
crear_excel_costos_kombucha()